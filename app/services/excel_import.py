"""기존 엑셀 요청서 데이터를 DB로 이관한다.

헤더 이름(별칭 포함)으로 컬럼을 찾는다:
- 팀(부서/팀명), 이름(성명), 계급(직급), 금액(충전금액/포인트),
  개인번호(번호/개인번호), 비고(메모/특이사항), 잔액(이월잔액/전월잔액, 선택)
"""

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import BalanceRecord, MonthlySnapshot, Person
from app.services.balance import compute_total, create_monthly_snapshot
from app.services.sync import RequestRow
from app.services.teams import get_or_create_team

HEADER_SYNONYMS: dict[str, list[str]] = {
    "team": ["팀", "부서", "팀명"],
    "name": ["이름", "성명"],
    "grade": ["계급", "직급"],
    "amount": ["금액", "충전", "충전금액", "포인트"],
    "personal_no": ["개인번호", "번호"],
    "note": ["비고", "메모", "특이사항"],
    "balance": ["잔액", "이월", "이월잔액", "전월잔액"],
}


@dataclass
class ImportRow:
    team: str
    name: str
    grade: str
    amount: int
    personal_no: str
    note: str
    balance: int = 0

    def to_request_row(self) -> RequestRow:
        return RequestRow(
            personal_no=self.personal_no,
            name=self.name,
            team=self.team,
            grade=self.grade,
            amount=self.amount,
            note=self.note,
        )


@dataclass
class ImportResult:
    month: str
    created_persons: int
    existing_persons: int
    records: int


def _find_header_index(row: tuple[object, ...], synonyms: list[str]) -> int | None:
    for cell in row:
        text = str(getattr(cell, "value", "") or "").strip()
        if any(text == syn or text.startswith(syn) for syn in synonyms):
            return list(row).index(cell)
    return None


def _cell_int(text: str) -> int:
    return int(text.replace(",", "")) if text.replace(",", "").isdigit() else 0


def _build_row(columns: dict[str, int], values: list[str]) -> ImportRow | None:
    def value(key: str, default: str = "") -> str:
        idx = columns.get(key)
        return values[idx] if idx is not None and idx < len(values) else default

    name = value("name")
    personal_no = value("personal_no")
    if not name or not personal_no:
        return None
    return ImportRow(
        team=value("team"),
        name=name,
        grade=value("grade"),
        amount=_cell_int(value("amount")),
        personal_no=personal_no,
        note=value("note"),
        balance=_cell_int(value("balance")),
    )


def read_rows(path: Path) -> list[ImportRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        return []
    rows: list[ImportRow] = []
    columns: dict[str, int] = {}

    for row in sheet.iter_rows():
        values = [str(c.value or "").strip() for c in row]
        if not any(values):
            continue
        if not columns:
            for key, synonyms in HEADER_SYNONYMS.items():
                idx = _find_header_index(row, synonyms)
                if idx is not None:
                    columns[key] = idx
            if "name" in columns and "personal_no" in columns:
                continue
            columns = {}
        else:
            row = _build_row(columns, values)
            if row is not None:
                rows.append(row)
    return rows


def import_excel(db: Session, path: Path, month: str) -> ImportResult:
    """엑셀 행을 인원·팀으로 이관하고 해당 월 스냅샷을 생성한다.

    기존 인원은 유지하고 누락된 인원만 추가하며, 월 스냅샷은 없을 때만 생성한다.
    """
    if db.scalar(select(MonthlySnapshot).where(MonthlySnapshot.month == month)) is not None:
        raise ValueError(f"{month} 월은 이미 처리되었습니다.")
    if db.scalar(select(Person).limit(1)) is not None:
        raise ValueError("DB에 이미 인원이 존재합니다. 빈 DB에서만 이관할 수 있습니다.")

    imported = read_rows(path)
    created = 0
    existing = 0
    for row in imported:
        person = db.scalar(
            select(Person).where(Person.personal_no == row.personal_no, Person.name == row.name)
        )
        if person is not None:
            existing += 1
            continue
        team = get_or_create_team(db, row.team) if row.team else None
        db.add(
            Person(
                personal_no=row.personal_no,
                name=row.name,
                grade=row.grade,
                team_id=team.id if team else None,
                status="active",
                current_carry_balance=row.balance,
                current_amount=row.amount,
            )
        )
        created += 1

    persons = (
        db.scalars(
            select(Person).where(Person.personal_no.in_([r.personal_no for r in imported]))
        ).all()
        if imported
        else []
    )
    records: list[BalanceRecord] = []
    for row in imported:
        person = next(
            (p for p in persons if p.personal_no == row.personal_no and p.name == row.name), None
        )
        if person is None:
            continue
        records.append(
            BalanceRecord(
                person_id=person.id,
                carry_balance=row.balance,
                amount=row.amount,
                usage=0,
                total=compute_total(row.amount, row.balance),
            )
        )
    if records:
        create_monthly_snapshot(db, month, records)
    return ImportResult(
        month=month, created_persons=created, existing_persons=existing, records=len(records)
    )
