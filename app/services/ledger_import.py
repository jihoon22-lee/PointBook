"""누적 포인트 장부의 마스킹된 검증과 이관용 데이터 모델.

이 모듈은 고정 형식의 ``간식비`` 시트에서 상세행 2:75만 읽는다. 원본의
76·77행 집계 수식은 신뢰하지 않으며, 공용 계정은 CJ:CL의 81:84에서 읽는다.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import delete, func, select
from sqlalchemy.orm import Session

from app.models import BalanceRecord, MonthlySnapshot, Person, Team
from app.services.identifiers import is_legacy_point_no, normalize_point_no


class LedgerImportError(ValueError):
    """개인정보를 포함하지 않는 누적 장부 검증 오류."""


@dataclass(frozen=True)
class LedgerAccount:
    point_no: str
    personal_no: str | None
    name: str
    grade: str
    team_name: str
    account_type: str
    status: str


@dataclass(frozen=True)
class LedgerRecord:
    point_no: str
    month: str
    carry_balance: int
    amount: int
    usage: int
    total: int


@dataclass(frozen=True)
class LedgerWarning:
    code: str
    cell: str


@dataclass(frozen=True)
class LedgerData:
    people: tuple[LedgerAccount, ...]
    shared_accounts: tuple[LedgerAccount, ...]
    records: tuple[LedgerRecord, ...]
    months: tuple[str, ...]
    warnings: tuple[LedgerWarning, ...]


@dataclass(frozen=True)
class LedgerMonthSummary:
    month: str
    count: int
    carry_balance: int
    amount: int
    usage: int
    total: int


@dataclass(frozen=True)
class LedgerImportPlan:
    data: LedgerData
    matched_account_ids: tuple[tuple[str, int], ...]
    delete_person_ids: tuple[int, ...]
    create_account_count: int
    update_account_count: int


@dataclass(frozen=True)
class LedgerApplyResult:
    accounts: int
    months: int
    records: int
    deleted_accounts: int


MONTH_BLOCKS: tuple[tuple[str, int], ...] = (
    ("2024-06", 7),
    ("2024-07", 10),
    ("2024-08", 13),
    ("2024-09", 16),
    ("2024-10", 19),
    ("2024-11", 22),
    ("2024-12", 25),
    ("2025-01", 28),
    ("2025-02", 31),
    ("2025-03", 34),
    ("2025-04", 37),
    ("2025-05", 40),
    ("2025-06", 43),
    ("2025-07", 46),
    ("2025-08", 49),
    ("2025-09", 52),
    ("2025-10", 55),
    ("2025-11", 58),
    ("2025-12", 61),
    ("2026-01", 64),
    ("2026-02", 67),
    ("2026-03", 70),
    ("2026-04", 73),
    ("2026-05", 76),
    ("2026-06", 79),
    ("2026-07", 82),
    ("2026-08", 85),
)
MONTHS: tuple[str, ...] = ("2024-05", *(month for month, _ in MONTH_BLOCKS))
EXPECTED_HEADERS = ("팀", "이름", "계급", "개인번호", "포인트번호")
SHARED_ROWS = range(81, 85)


def _cell(column: int, row: int = 1) -> str:
    return f"{get_column_letter(column)}{row}"


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _money(value: object, coordinate: str) -> int:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise LedgerImportError(f"{coordinate}: 금액이 비어 있거나 숫자가 아닙니다.")
    if int(value) != value:
        raise LedgerImportError(f"{coordinate}: 금액은 정수여야 합니다.")
    return int(value)


def _previous_year_month(month: str) -> tuple[int, int]:
    year, month_no = (int(part) for part in month.split("-"))
    if month_no == 1:
        return year - 1, 12
    return year, month_no - 1


def _validate_headers(ws: Any) -> tuple[LedgerWarning, ...]:
    for column, expected in enumerate(EXPECTED_HEADERS, 1):
        coordinate = _cell(column)
        if _text(ws[coordinate].value) != expected:
            raise LedgerImportError(f"{coordinate}: 기본 헤더가 올바르지 않습니다.")
    if _text(ws["F1"].value) != "5월분":
        raise LedgerImportError("F1: 최초 지급 월 헤더가 올바르지 않습니다.")

    warnings: list[LedgerWarning] = []
    for month, start in MONTH_BLOCKS:
        carry_cell = _cell(start)
        amount_cell = _cell(start + 1)
        total_cell = _cell(start + 2)
        carry_header = _text(ws[carry_cell].value)
        carry_match = re.search(r"(\d{2})년\s*0?(\d{1,2})월\s*잔액", carry_header)
        previous_year, previous_month = _previous_year_month(month)
        if carry_match is None or (
            int(carry_match.group(1)) != previous_year % 100
            or int(carry_match.group(2)) != previous_month
        ):
            raise LedgerImportError(f"{carry_cell}: 잔액 월 헤더가 열 매핑과 다릅니다.")

        amount_header = _text(ws[amount_cell].value)
        amount_match = re.search(r"(\d{1,2})월분", amount_header)
        expected_month = int(month[-2:])
        if amount_match is None:
            raise LedgerImportError(f"{amount_cell}: 지급 월 헤더를 읽을 수 없습니다.")
        if int(amount_match.group(1)) != expected_month:
            warnings.append(LedgerWarning(code="PAYMENT_MONTH_LABEL_MISMATCH", cell=amount_cell))
        if not _text(ws[total_cell].value).startswith("합계"):
            raise LedgerImportError(f"{total_cell}: 합계 헤더가 올바르지 않습니다.")
    return tuple(warnings)


def _parse_point_no(value: object, coordinate: str) -> str:
    try:
        return normalize_point_no(_text(value))
    except ValueError as exc:
        raise LedgerImportError(f"{coordinate}: 포인트번호 형식이 올바르지 않습니다.") from exc


def _parse_people(ws: Any) -> tuple[LedgerAccount, ...]:
    people: list[LedgerAccount] = []
    seen: set[str] = set()
    for row in range(2, 76):
        team_name = _text(ws.cell(row, 1).value)
        name = _text(ws.cell(row, 2).value)
        grade = _text(ws.cell(row, 3).value)
        personal_no = _text(ws.cell(row, 4).value)
        point_no = _parse_point_no(ws.cell(row, 5).value, f"E{row}")
        if not all((team_name, name, grade, personal_no)):
            raise LedgerImportError(f"A{row}: 일반 인원 기본 정보가 비어 있습니다.")
        if point_no in seen:
            raise LedgerImportError(f"E{row}: 중복 포인트번호가 있습니다.")
        seen.add(point_no)
        latest_amount = ws.cell(row, 86).value  # CH
        active = (
            isinstance(latest_amount, (int, float))
            and not isinstance(latest_amount, bool)
            and latest_amount > 0
        )
        people.append(
            LedgerAccount(
                point_no=point_no,
                personal_no=personal_no,
                name=name,
                grade=grade,
                team_name=team_name,
                account_type="person",
                status="active" if active else "inactive",
            )
        )
    return tuple(people)


def _parse_shared_accounts(ws: Any, person_numbers: set[str]) -> tuple[LedgerAccount, ...]:
    accounts: list[LedgerAccount] = []
    seen = set(person_numbers)
    for row in SHARED_ROWS:
        name = _text(ws.cell(row, 88).value)  # CJ
        team_value = ws.cell(row, 89).value  # CK
        point_no = _parse_point_no(ws.cell(row, 90).value, f"CL{row}")
        if not name or not isinstance(team_value, (int, float)) or isinstance(team_value, bool):
            raise LedgerImportError(f"CJ{row}: 공용 계정 정보가 올바르지 않습니다.")
        if point_no in seen:
            raise LedgerImportError(f"CL{row}: 중복 포인트번호가 있습니다.")
        seen.add(point_no)
        team_number = int(team_value)
        accounts.append(
            LedgerAccount(
                point_no=point_no,
                personal_no=None,
                name=name,
                grade="",
                team_name=f"{team_number}팀" if team_number else "",
                account_type="shared",
                status="active",
            )
        )
    return tuple(accounts)


def _parse_records(ws: Any, people: tuple[LedgerAccount, ...]) -> tuple[LedgerRecord, ...]:
    records: list[LedgerRecord] = []
    for offset, person in enumerate(people, 2):
        previous_total: int | None = None
        initial_value = ws.cell(offset, 6).value  # F
        if initial_value is not None:
            amount = _money(initial_value, f"F{offset}")
            records.append(
                LedgerRecord(
                    point_no=person.point_no,
                    month="2024-05",
                    carry_balance=0,
                    amount=amount,
                    usage=0,
                    total=amount,
                )
            )
            previous_total = amount

        for month, start in MONTH_BLOCKS:
            values = [ws.cell(offset, column).value for column in range(start, start + 3)]
            # 원본 합계 열은 동적 배열 수식이라 미등록 행도 캐시값 0을 가진다.
            # 이월·지급이 모두 비고 합계가 비거나 0이면 실제로는 미등록 월이다.
            if values[0] is None and values[1] is None and values[2] in (None, 0):
                continue
            carry_cell, amount_cell, total_cell = (
                _cell(start + delta, offset) for delta in range(3)
            )
            carry = _money(values[0], carry_cell)
            amount = 0 if values[1] is None else _money(values[1], amount_cell)
            total = _money(values[2], total_cell)
            if total != carry + amount:
                raise LedgerImportError(f"{total_cell}: 합계가 이월 잔액과 지급액의 합과 다릅니다.")
            usage = 0 if previous_total is None else previous_total - carry
            records.append(
                LedgerRecord(
                    point_no=person.point_no,
                    month=month,
                    carry_balance=carry,
                    amount=amount,
                    usage=usage,
                    total=total,
                )
            )
            previous_total = total
    return tuple(records)


def parse_ledger(path: Path) -> LedgerData:
    """누적 장부를 읽고 개인정보를 출력하지 않는 오류로 형식과 수식을 검증한다."""
    try:
        # 고정 좌표를 반복 조회하므로 일반 모드가 read-only 모드의 반복 XML 탐색보다 빠르다.
        workbook = load_workbook(path, data_only=True, read_only=False)
    except Exception as exc:
        raise LedgerImportError("장부 파일을 열 수 없습니다.") from exc
    if "간식비" not in workbook.sheetnames:
        workbook.close()
        raise LedgerImportError("필수 '간식비' 시트가 없습니다.")
    try:
        ws = workbook["간식비"]
        warnings = _validate_headers(ws)
        people = _parse_people(ws)
        shared_accounts = _parse_shared_accounts(ws, {person.point_no for person in people})
        records = _parse_records(ws, people)
        return LedgerData(
            people=people,
            shared_accounts=shared_accounts,
            records=records,
            months=MONTHS,
            warnings=warnings,
        )
    finally:
        workbook.close()


def month_summary(data: LedgerData, month: str) -> LedgerMonthSummary:
    records = [record for record in data.records if record.month == month]
    return LedgerMonthSummary(
        month=month,
        count=len(records),
        carry_balance=sum(record.carry_balance for record in records),
        amount=sum(record.amount for record in records),
        usage=sum(record.usage for record in records),
        total=sum(record.total for record in records),
    )


def validate_expected_totals(data: LedgerData) -> None:
    """승인된 실제 장부의 고정 불변식을 검증한다."""
    active = sum(account.status == "active" for account in data.people)
    inactive = sum(account.status == "inactive" for account in data.people)
    negative_usage = sum(record.usage < 0 for record in data.records)
    first = month_summary(data, "2024-05")
    last = month_summary(data, "2026-08")
    checks = (
        (len(data.people), 74, "일반 인원 수"),
        (len(data.shared_accounts), 4, "공용 계정 수"),
        (len(data.months), 28, "월 수"),
        (data.months[0], "2024-05", "최초 월"),
        (data.months[-1], "2026-08", "최종 월"),
        (len(data.records), 1392, "잔액 기록 수"),
        (active, 44, "active 일반 인원 수"),
        (inactive, 30, "inactive 일반 인원 수"),
        (negative_usage, 5, "음수 순사용 수"),
        (first.count, 33, "최초 월 기록 수"),
        (first.amount, 1_385_000, "최초 월 지급액"),
        (first.total, 1_385_000, "최초 월 총잔액"),
        (last.count, 74, "최종 월 기록 수"),
        (last.carry_balance, 2_848_212, "최종 월 이월 잔액"),
        (last.amount, 1_716_000, "최종 월 지급액"),
        (last.total, 4_564_212, "최종 월 총잔액"),
    )
    for actual_value, expected_value, label in checks:
        if actual_value != expected_value:
            raise LedgerImportError(f"예상 검증값 불일치: {label}")


def analyze_ledger_import(
    db: Session,
    data: LedgerData,
    *,
    replace_empty_history_people: bool = False,
) -> LedgerImportPlan:
    """DB를 변경하지 않고 누적 장부 적용 계획을 만든다."""
    snapshot_count = db.scalar(select(func.count(MonthlySnapshot.id))) or 0
    record_count = db.scalar(select(func.count(BalanceRecord.id))) or 0
    if snapshot_count or record_count:
        raise LedgerImportError("기존 월별 이력이 있어 누적 장부를 적용할 수 없습니다.")

    source_accounts = (*data.people, *data.shared_accounts)
    existing = list(db.scalars(select(Person).order_by(Person.id)).all())
    existing_by_point = {person.point_no: person for person in existing}
    matched: dict[str, int] = {}
    matched_ids: set[int] = set()

    for account in source_accounts:
        direct = existing_by_point.get(account.point_no)
        if direct is not None:
            matched[account.point_no] = direct.id
            matched_ids.add(direct.id)
            continue
        if account.account_type != "person":
            continue
        candidates = [
            person
            for person in existing
            if person.id not in matched_ids
            and is_legacy_point_no(person.point_no)
            and person.personal_no == account.personal_no
            and person.name == account.name
        ]
        if len(candidates) > 1:
            raise LedgerImportError("기존 계정 백필에 다중 후보가 있습니다.")
        if len(candidates) == 1:
            matched[account.point_no] = candidates[0].id
            matched_ids.add(candidates[0].id)

    unmatched = [person for person in existing if person.id not in matched_ids]
    if unmatched and not replace_empty_history_people:
        raise LedgerImportError(
            f"기존 미매칭 계정 {len(unmatched)}개가 있습니다. 빈 이력 계정 교체 옵션이 필요합니다."
        )
    if unmatched and replace_empty_history_people and (len(existing) != 2 or len(unmatched) != 2):
        raise LedgerImportError(
            "빈 이력 계정 교체는 기존 계정 전체가 미매칭 테스트 계정 정확히 2개일 때만 가능합니다."
        )
    delete_ids = tuple(person.id for person in unmatched) if replace_empty_history_people else ()
    return LedgerImportPlan(
        data=data,
        matched_account_ids=tuple(sorted(matched.items())),
        delete_person_ids=delete_ids,
        create_account_count=len(source_accounts) - len(matched),
        update_account_count=len(matched),
    )


def _team_map(db: Session, data: LedgerData) -> dict[str, Team]:
    teams = {team.name: team for team in db.scalars(select(Team)).all()}
    required_names = {
        account.team_name for account in (*data.people, *data.shared_accounts) if account.team_name
    }
    for name in sorted(required_names):
        if name not in teams:
            team = Team(name=name)
            db.add(team)
            teams[name] = team
    db.flush()
    return teams


def _validate_database_invariants(db: Session, data: LedgerData) -> None:
    source_accounts = (*data.people, *data.shared_accounts)
    people = list(db.scalars(select(Person)).all())
    snapshots = list(db.scalars(select(MonthlySnapshot)).all())
    records = list(db.scalars(select(BalanceRecord)).all())
    checks = (
        (len(people), len(source_accounts), "계정 수"),
        (len(snapshots), len(data.months), "월 수"),
        (len(records), len(data.records), "잔액 기록 수"),
        (
            sum(person.account_type == "person" for person in people),
            len(data.people),
            "일반 인원 수",
        ),
        (
            sum(person.account_type == "shared" for person in people),
            len(data.shared_accounts),
            "공용 계정 수",
        ),
        (
            sum(record.usage < 0 for record in records),
            sum(r.usage < 0 for r in data.records),
            "음수 순사용 수",
        ),
    )
    for actual, expected, label in checks:
        if actual != expected:
            raise LedgerImportError(f"DB 사후 검증 실패: {label}")
    if {snapshot.month for snapshot in snapshots} != set(data.months):
        raise LedgerImportError("DB 사후 검증 실패: 월 범위")
    if len({person.point_no for person in people}) != len(people) or any(
        re.fullmatch(r"\d{8}", person.point_no) is None for person in people
    ):
        raise LedgerImportError("DB 사후 검증 실패: 포인트번호")

    expected_by_month = {month: month_summary(data, month) for month in data.months}
    snapshot_by_id = {snapshot.id: snapshot.month for snapshot in snapshots}
    for month, expected_summary in expected_by_month.items():
        monthly_records = [
            record for record in records if snapshot_by_id[record.snapshot_id] == month
        ]
        actual_aggregate = (
            len(monthly_records),
            sum(record.carry_balance for record in monthly_records),
            sum(record.amount for record in monthly_records),
            sum(record.usage for record in monthly_records),
            sum(record.total for record in monthly_records),
        )
        wanted_aggregate = (
            expected_summary.count,
            expected_summary.carry_balance,
            expected_summary.amount,
            expected_summary.usage,
            expected_summary.total,
        )
        if actual_aggregate != wanted_aggregate:
            raise LedgerImportError("DB 사후 검증 실패: 월별 합계")


def apply_ledger_import(db: Session, plan: LedgerImportPlan) -> LedgerApplyResult:
    """계획을 현재 트랜잭션에 적용한다. 성공 후 commit은 호출자가 수행한다."""
    try:
        if (db.scalar(select(func.count(MonthlySnapshot.id))) or 0) or (
            db.scalar(select(func.count(BalanceRecord.id))) or 0
        ):
            raise LedgerImportError("기존 월별 이력이 있어 누적 장부를 적용할 수 없습니다.")
        if plan.delete_person_ids:
            db.execute(delete(Person).where(Person.id.in_(plan.delete_person_ids)))

        teams = _team_map(db, plan.data)
        matched = dict(plan.matched_account_ids)
        source_accounts = (*plan.data.people, *plan.data.shared_accounts)
        account_by_point: dict[str, Person] = {}
        for account in source_accounts:
            person_id = matched.get(account.point_no)
            person = db.get(Person, person_id) if person_id is not None else None
            if person is None:
                person = Person(point_no=account.point_no, name=account.name)
                db.add(person)
            person.point_no = account.point_no
            person.personal_no = account.personal_no
            person.name = account.name
            person.grade = account.grade
            person.status = account.status
            person.account_type = account.account_type
            person.team_id = teams[account.team_name].id if account.team_name else None
            person.current_carry_balance = 0
            person.current_amount = 0
            account_by_point[account.point_no] = person
        db.flush()

        snapshots: dict[str, MonthlySnapshot] = {}
        for month in plan.data.months:
            snapshot = MonthlySnapshot(month=month)
            db.add(snapshot)
            snapshots[month] = snapshot
        db.flush()

        latest: dict[str, LedgerRecord] = {}
        for source in plan.data.records:
            person = account_by_point[source.point_no]
            db.add(
                BalanceRecord(
                    snapshot_id=snapshots[source.month].id,
                    person_id=person.id,
                    carry_balance=source.carry_balance,
                    amount=source.amount,
                    usage=source.usage,
                    total=source.total,
                )
            )
            latest[source.point_no] = source
        for point_no, source in latest.items():
            person = account_by_point[point_no]
            person.current_carry_balance = source.carry_balance
            person.current_amount = source.amount
        db.flush()
        _validate_database_invariants(db, plan.data)
        return LedgerApplyResult(
            accounts=len(source_accounts),
            months=len(plan.data.months),
            records=len(plan.data.records),
            deleted_accounts=len(plan.delete_person_ids),
        )
    except Exception:
        db.rollback()
        raise
