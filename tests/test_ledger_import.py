from pathlib import Path
from shutil import copyfile

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select

from app.models import BalanceRecord, MonthlySnapshot, Person
from app.services.ledger_import import (
    LedgerImportError,
    analyze_ledger_import,
    apply_ledger_import,
    parse_ledger,
)

MONTH_BLOCKS = (
    ("2024-06", 7),
    ("2024-07", 10),
    ("2024-08", 13),
    ("2024-09", 16),
    ("2024-10", 19),
    ("2024-11", 22),
    ("2024-12", 25),
    ("2025-01", 28),
    ("2025-02", 31),
    ("2025-03", 34),
    ("2025-04", 37),
    ("2025-05", 40),
    ("2025-06", 43),
    ("2025-07", 46),
    ("2025-08", 49),
    ("2025-09", 52),
    ("2025-10", 55),
    ("2025-11", 58),
    ("2025-12", 61),
    ("2026-01", 64),
    ("2026-02", 67),
    ("2026-03", 70),
    ("2026-04", 73),
    ("2026-05", 76),
    ("2026-06", 79),
    ("2026-07", 82),
    ("2026-08", 85),
)


def _build_ledger(path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "간식비"
    for column, value in enumerate(("팀", "이름", "계급", "개인번호", "포인트번호"), 1):
        ws.cell(1, column, value)
    ws["F1"] = "5월분"
    for month, start in MONTH_BLOCKS:
        year, month_no = (int(part) for part in month.split("-"))
        previous_month = 12 if month_no == 1 else month_no - 1
        previous_year = year - 1 if month_no == 1 else year
        ws.cell(1, start, f"{previous_year % 100}년 {previous_month}월 잔액")
        displayed_month = month_no
        if month in {
            "2026-03",
            "2026-04",
            "2026-05",
            "2026-06",
            "2026-07",
            "2026-08",
        }:
            displayed_month -= 1
        ws.cell(1, start + 1, f"{displayed_month}월분")
        ws.cell(1, start + 2, "합계")

    for index, row in enumerate(range(2, 76), 1):
        ws.cell(row, 1, f"{(index - 1) % 3 + 1}팀")
        ws.cell(row, 2, f"합성인원{index:02d}")
        ws.cell(row, 3, "합성계급")
        ws.cell(row, 4, f"P{index:04d}")
        ws.cell(row, 5, f"{index:08d}")

    # 최초 기록, 음수 순사용, 중간 공백 뒤 복귀를 한 번씩 포함한다.
    ws["F2"] = 100
    ws["G2"], ws["H2"], ws["I2"] = 80, 20, 100
    ws["J2"], ws["K2"], ws["L2"] = 150, 0, 150
    ws["F3"] = 100
    ws["J3"], ws["K3"], ws["L3"] = 40, 10, 50
    ws["I4"] = 0  # 원본 동적 배열 합계 수식의 미등록 행 캐시값

    # 집계행은 파서가 절대로 읽지 않아야 한다.
    for column in range(6, 88):
        ws.cell(76, column, 999_999_999)
        ws.cell(77, column, 888_888_888)

    for offset, (name, team_number) in enumerate(
        (("1팀 공용", 1), ("2팀 공용", 2), ("3팀 공용", 3), ("소방서 공용", 0)),
        81,
    ):
        ws.cell(offset, 88, name)
        ws.cell(offset, 89, team_number)
        ws.cell(offset, 90, f"{90 + offset:08d}")
    wb.save(path)
    return path


@pytest.fixture(scope="module")
def base_ledger_path(tmp_path_factory: pytest.TempPathFactory) -> Path:
    return _build_ledger(tmp_path_factory.mktemp("ledger-base") / "synthetic-ledger.xlsx")


@pytest.fixture
def ledger_path(tmp_path: Path, base_ledger_path: Path) -> Path:
    return Path(copyfile(base_ledger_path, tmp_path / "synthetic-ledger.xlsx"))


def test_parse_ledger_has_explicit_28_month_mapping(ledger_path):
    data = parse_ledger(ledger_path)
    assert data.months[0] == "2024-05"
    assert data.months[-1] == "2026-08"
    assert len(data.months) == 28
    assert len(data.people) == 74
    assert len(data.shared_accounts) == 4


def test_parse_ledger_blank_months_and_ignored_totals(ledger_path):
    data = parse_ledger(ledger_path)
    assert len(data.records) == 5
    assert max(record.total for record in data.records) < 999_999_999
    assert not any(
        record.month == "2024-06" and record.point_no == "00000002" for record in data.records
    )


def test_parse_ledger_signed_usage_and_latest_existing_record(ledger_path):
    data = parse_ledger(ledger_path)
    by_key = {(record.point_no, record.month): record for record in data.records}
    assert by_key[("00000001", "2024-07")].usage == -50
    assert by_key[("00000002", "2024-07")].usage == 60


def test_parse_ledger_reports_known_late_header_warnings(ledger_path):
    data = parse_ledger(ledger_path)
    assert [warning.cell for warning in data.warnings] == ["BS1", "BV1", "BY1", "CB1", "CE1", "CH1"]


@pytest.mark.parametrize("cell", ["G2", "I2"])
def test_parse_ledger_rejects_missing_required_balance_cell(ledger_path, cell):
    wb = load_workbook(ledger_path)
    wb["간식비"][cell] = None
    wb.save(ledger_path)
    with pytest.raises(LedgerImportError, match=cell):
        parse_ledger(ledger_path)


def test_parse_ledger_rejects_total_mismatch_without_identity(ledger_path):
    wb = load_workbook(ledger_path)
    wb["간식비"]["I2"] = 999
    wb.save(ledger_path)
    with pytest.raises(LedgerImportError, match="I2") as caught:
        parse_ledger(ledger_path)
    assert "합성인원" not in str(caught.value)
    assert "00000001" not in str(caught.value)


def test_parse_ledger_shared_accounts_have_no_records(ledger_path):
    data = parse_ledger(ledger_path)
    shared_numbers = {account.point_no for account in data.shared_accounts}
    assert shared_numbers.isdisjoint(record.point_no for record in data.records)
    assert all(account.status == "active" for account in data.shared_accounts)


def test_parse_ledger_rejects_duplicate_point_numbers(ledger_path):
    wb = load_workbook(ledger_path)
    wb["간식비"]["E3"] = wb["간식비"]["E2"].value
    wb.save(ledger_path)
    with pytest.raises(LedgerImportError, match="중복"):
        parse_ledger(ledger_path)


def test_fixture_declares_every_mapped_column(ledger_path):
    wb = load_workbook(ledger_path, read_only=True)
    ws = wb["간식비"]
    assert get_column_letter(MONTH_BLOCKS[-1][1] + 2) == "CI"
    assert ws["CI1"].value == "합계"


def _counts(db):
    return (
        db.scalar(select(func.count(Person.id))) or 0,
        db.scalar(select(func.count(MonthlySnapshot.id))) or 0,
        db.scalar(select(func.count(BalanceRecord.id))) or 0,
    )


def test_analyze_ledger_import_is_read_only(client, db, ledger_path):
    data = parse_ledger(ledger_path)
    before = _counts(db)
    plan = analyze_ledger_import(db, data)
    assert plan.create_account_count == 78
    assert plan.update_account_count == 0
    assert _counts(db) == before


def test_apply_ledger_import_creates_accounts_months_and_records(client, db, ledger_path):
    data = parse_ledger(ledger_path)
    plan = analyze_ledger_import(db, data)
    result = apply_ledger_import(db, plan)
    db.commit()
    assert result.accounts == 78
    assert _counts(db) == (78, 28, 5)
    assert db.scalar(select(func.count(Person.id)).where(Person.account_type == "shared")) == 4


def test_apply_updates_one_exact_legacy_match_in_place(client, db, ledger_path):
    data = parse_ledger(ledger_path)
    source = data.people[0]
    legacy = Person(
        point_no="L0000001",
        personal_no=source.personal_no,
        name=source.name,
        grade="",
        status="active",
        account_type="person",
    )
    db.add(legacy)
    db.commit()
    legacy_id = legacy.id
    plan = analyze_ledger_import(db, data)
    assert plan.update_account_count == 1
    apply_ledger_import(db, plan)
    db.commit()
    updated = db.get(Person, legacy_id)
    assert updated is not None
    assert updated.point_no == source.point_no


def test_analyze_rejects_ambiguous_legacy_match(client, db, ledger_path):
    data = parse_ledger(ledger_path)
    source = data.people[0]
    db.add_all(
        [
            Person(
                point_no=f"L{index:07d}",
                personal_no=source.personal_no,
                name=source.name,
                grade="",
                status="active",
                account_type="person",
            )
            for index in (1, 2)
        ]
    )
    db.commit()
    with pytest.raises(LedgerImportError, match="다중 후보") as caught:
        analyze_ledger_import(db, data, replace_empty_history_people=True)
    assert source.name not in str(caught.value)


def test_analyze_replacement_requires_empty_history(client, db, ledger_path):
    data = parse_ledger(ledger_path)
    extra = Person(
        point_no="L9999999",
        personal_no="legacy",
        name="기존 테스트",
        grade="",
        status="active",
        account_type="person",
    )
    db.add(extra)
    db.flush()
    snapshot = MonthlySnapshot(month="2030-01")
    db.add(snapshot)
    db.flush()
    db.add(
        BalanceRecord(
            snapshot_id=snapshot.id,
            person_id=extra.id,
            carry_balance=0,
            amount=0,
            usage=0,
            total=0,
        )
    )
    db.commit()
    with pytest.raises(LedgerImportError, match="이력"):
        analyze_ledger_import(db, data, replace_empty_history_people=True)


@pytest.mark.parametrize("existing_count", [1, 3])
def test_analyze_replacement_requires_exactly_two_unmatched_accounts(
    client, db, ledger_path, existing_count
):
    data = parse_ledger(ledger_path)
    db.add_all(
        [
            Person(
                point_no=f"L{index:07d}",
                personal_no=f"legacy-{index}",
                name=f"교체 테스트 {index}",
                grade="",
                status="active",
                account_type="person",
            )
            for index in range(1, existing_count + 1)
        ]
    )
    db.commit()

    with pytest.raises(LedgerImportError, match="정확히 2개"):
        analyze_ledger_import(db, data, replace_empty_history_people=True)


def test_analyze_replacement_accepts_exactly_two_unmatched_accounts(client, db, ledger_path):
    data = parse_ledger(ledger_path)
    db.add_all(
        [
            Person(
                point_no=f"L{index:07d}",
                personal_no=f"legacy-{index}",
                name=f"교체 테스트 {index}",
                grade="",
                status="active",
                account_type="person",
            )
            for index in (1, 2)
        ]
    )
    db.commit()

    plan = analyze_ledger_import(db, data, replace_empty_history_people=True)

    assert len(plan.delete_person_ids) == 2


def test_apply_rolls_back_everything_on_invariant_failure(client, db, ledger_path, monkeypatch):
    from app.services import ledger_import

    data = parse_ledger(ledger_path)
    plan = analyze_ledger_import(db, data)

    def fail(*args, **kwargs):
        raise LedgerImportError("사후 검증 실패")

    monkeypatch.setattr(ledger_import, "_validate_database_invariants", fail)
    with pytest.raises(LedgerImportError, match="사후 검증"):
        apply_ledger_import(db, plan)
    assert _counts(db) == (0, 0, 0)


def test_second_apply_is_rejected(client, db, ledger_path):
    data = parse_ledger(ledger_path)
    apply_ledger_import(db, analyze_ledger_import(db, data))
    db.commit()
    with pytest.raises(LedgerImportError, match="월별 이력"):
        analyze_ledger_import(db, data)
